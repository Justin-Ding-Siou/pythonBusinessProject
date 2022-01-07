# For each single excel file, create one CSV file per sheet. 
# The filenames of the CSV files should be <excel filename>_<sheet title>.csv,

import os
import csv
import openpyxl

for filename in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not filename.endswith('.xlsx'):
        continue # skip non-csv files
    
    fname, ext = os.path.splitext(filename) 
    
    # open file using openpyxl
    wb = openpyxl.load_workbook(filename)

    for sh in wb.sheetnames:
    # Loop through every sheet in the workbook.
        sheet = wb[sh]
        # Create the CSV filename from the Excel filename and sheet title.
        # Create the csv.writer object for this CSV file.
        csv_fname = fname + '_'  + sh + '.csv'
        print(csv_fname)
        cf = open(csv_fname, 'w', newline='')
        csv_writer = csv.writer(cf)
        
        # Loop through every row in the sheet.
        for rowdata in sheet.iter_rows(min_row=1, values_only=True):
            csv_writer.writerow(rowdata)
        
        cf.close()